import pyseq
import subprocess
import json
import re
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def get_latest_xlsx(app_instance):
    """
    Get path of latest version excel file in selected directory.
    Returns v001 if no file exists in the selected directory.

    Returns:
        str or None: Path of excel file
    """
    date_directory_path = app_instance.xl_manager.get_date_path()
    versioned_files = get_xl_files(app_instance, date_directory_path)
    if not versioned_files:
        meta_data_list = export_metadata(app_instance, date_directory_path)
        # If there is no xlsx file. Make _v001.xlsx
        print("[INFO] No versioned file found. v001 will be created")
        new_xlsx_path = save_list_as_xlsx(app_instance, meta_data_list, date_directory_path)
        latest_xlsx_path = new_xlsx_path
    else:
        latest_version_tuple = max(versioned_files)
        latest_xlsx_path = os.path.join(date_directory_path, latest_version_tuple[1])
    return latest_xlsx_path
    
def export_metadata(app_instance, date_path):
    """
    Extract metadata for each shot in the selected directory

    Args:
        date_path: Selected date directory path by user
    
    Returns:
        list: List of meta data dictionary
        list ex: [{meta data of shot1}, {meta data of shot2}, ... ,{meta data of shotN}]
    """
    meta_list = []
    thumbnails_dir = os.path.join(date_path, "thumbnails")
    os.makedirs(thumbnails_dir, exist_ok=True)

    # .../scan/{date} 순회 하면서 Sequence 객체 get
    for root, dirs, seqs in pyseq.walk(date_path):
        dirs[:] = [d for d in dirs if d != "thumbnails"]
        first_frame_path = "" #exr의 첫 프레임 path / mov path
        for seq in seqs:
            scan_name = seq.head().rstrip(".")
            _, ext = os.path.splitext(seq.name)
            # EXR sequnece 폴더의 첫 프레임을 thumnail(JPG)로 변환
            if ext == ".exr":
                first_frame_path = seq[0].path
                thumb_name = scan_name + ".jpg"
                thumb_path = os.path.join(thumbnails_dir, thumb_name)
                # If thumbnail not exist, make thumbnail
                if not os.path.exists(thumb_path):
                    make_excel_thumbnail_result = make_excel_thumbnail(first_frame_path, thumb_path)
                    if not make_excel_thumbnail_result:
                        error_msg = f"Error occurred while attempting to convert thumbnail.\nIO Manager skips the thumbnail creation process."
                        app_instance.show_error_dialog(error_msg)
                        thumb_path = ""
                # Skip making thumbnail if thumbnail exist
                else:
                    print(f"[SKIP] Thumbnail already exist: {thumb_path}")
            # MOV 폴더의 첫 프레임 thumbnial(JPG) 추출
            elif ext == ".mov":
                # mov를 shot 단위의 exr로 만들고 이후 과정은 똑같이
                # mov_path = os.path.join(root, seq.name)
                # scan_data_path = mov_path # for export meta data
                # thumb_name = os.path.splitext(seq.name)[0] + ".jpg"
                # thumb_path = os.path.join(thumbnails_dir, thumb_name)
                # if not os.path.exists(thumb_path):
                #     if mov_to_jpg(scan_data_path, thumb_path):
                #         print(f"[OK] Thumbnail created: {thumb_path}")
                #     else:
                #         print(f"[FAIL] Thumbnail create failed: {scan_data_path}")
                #         thumb_path = ""

                # else:
                #     print(f"[SKIP] Thumbnail already exist: {thumb_path}")
                continue
            else:
                print(f"[SKIP] {seq} is not EXR OR MOV")
                continue

            # Exporting metadata using EXIF tool 
            result = subprocess.run(
                ["exiftool", "-json", first_frame_path],
                capture_output=True, 
                text=True
            )
            # meta data json 파싱
            # ex) result(string) -> meta(json) 
            meta = json.loads(result.stdout)[0]
            meta["scan name"] = scan_name
            meta["thumbnail"] = thumb_path
            meta_list.append(meta)
    return meta_list

def make_excel_thumbnail(input_exr_path, output_jpg_path):
    """
    Convert EXR to JPG
    Args:
        input_exr_path (str): Path of first frame of exr sequences
        output_jpg_path (str): Path of destination of thumbnail image

    Returns:
        bool: Conversion successful

    """
    cmd = [
        "oiiotool",
        input_exr_path,
        "--colorconvert", "ACES", "sRGB",
        "-o", output_jpg_path
    ]

    result = subprocess.run(cmd, check=False)
    if result.returncode != 0:
        print(f"[ERROR] Error occurred while attempting to convert thumbnail")
        return False
    print(f"[COMPLETE] {input_exr_path} >>>>> {output_jpg_path}")
    return True

def get_xl_files(app_instance, date_directory_path):
    """
    Get excel files in selected date directory

    Returns: 
        list: Tuple list consisting of version, file name
    """
    versioned_files = []
    if not date_directory_path:
        app_instance.show_error_dialog(f"Please select date directory first")
        return versioned_files
    # ex ) date_prefix: 20250529
    date_prefix = os.path.basename(date_directory_path)
    pattern = re.compile(rf"{date_prefix}_list_v(\d{{3}})\.xlsx$")
    for fname in os.listdir(date_directory_path):
        match = pattern.match(fname)
        if match:
            version = int(match.group(1))
            versioned_files.append((version, fname))
    # ex ) versioned_files: [(1, 20250529_list_v001.xlsx), (2, 20250529_list_v002.xlsx), ..., (14, 20250529_list_v014.xlsx)]
    return versioned_files

def save_list_as_xlsx(app_instance, meta_data_list, date_directory_path, xl_name = None):
    """
    Save meta data list as xlsx

    Args:
        List of meta data dictionary
        meta_data_list ex => [{meta_data1}, {meta_data2}, ... , {meta_dataN}]
    
    Returns:
        str or None : Path of saved excel file
    """
    if not meta_data_list:
        error_msg = f"No shot for exporting excel file"
        app_instance.show_error_dialog(error_msg)
        return None

    if not xl_name:
        xl_name = f"{os.path.basename(date_directory_path)}_list_v001.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    default_fields = ["thumbnail", "shot name", "seq name", "scan name"]
    all_keys_set = set()
    for meta_data in meta_data_list:
        for key in meta_data.keys():
            all_keys_set.add(key)

    all_keys_set -= set(default_fields)

    # append headers
    all_fields = default_fields + list(all_keys_set) 
    ws.append(all_fields)

    # Insert metadata into the worksheet by iterating list of dictionaries
    for row_idx, meta_data in enumerate(meta_data_list, start=2): # Start at row2(Row1 : header)
        for col_idx, field in enumerate(all_fields, start=1):
            value = meta_data.get(field, "")
            # if value's type is list
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            ws.cell(row=row_idx, column=col_idx, value=value)

        # Insert thumbnail to cell
        thumb_path = meta_data.get("thumbnail")
        if thumb_path and os.path.exists(thumb_path):
            col_letter = get_column_letter(all_fields.index("thumbnail") + 1)
            cell_ref = f"{col_letter}{row_idx}"
            img = Image(thumb_path)
            img.width = 192
            img.height = 108
            ws.add_image(img, cell_ref)
    xl_path = os.path.join(date_directory_path, xl_name)
    wb.save(xl_path)

    print(f"[COMPLETE] Metadata exported to:\n{xl_path}")
    return xl_path

def get_new_xlsx(app_instance):
    """
    Get (latest_version + 1) xlsx file path

    Returns:
        str: new xlsx file path
    """
    date_directory_path = app_instance.xl_manager.get_date_path()
    versioned_files = get_xl_files(app_instance, date_directory_path)
    # No selected date path 
    if not versioned_files:
        new_version_xl = ""
        return new_version_xl
    latest_version_tuple = max(versioned_files)
    next_version = latest_version_tuple[0] + 1
    date_dir_name = os.path.basename(date_directory_path)
    file_name = f"{date_dir_name}_list_v{next_version:03d}.xlsx"
    new_version_xl = os.path.join(date_directory_path, file_name)
    return new_version_xl

def save_table_to_xlsx(app_instance, save_path):
    """
    Save table widget as xlsx
    Table -> dictionary list -> xlsx

    Returns:
        str : Path of saved excel file
    """
    date_directory_path = app_instance.xl_manager.get_date_path()
    meta_list = table_to_meta_list(app_instance)        
    for meta in meta_list:
        scan_name = meta.get("scan name")
        if scan_name:
            thumb_path = os.path.join(date_directory_path, "thumbnails", scan_name + ".jpg")
            if os.path.exists(thumb_path):
                meta["thumbnail"] = thumb_path
            else:
                meta["thumbnail"] = ""

    save_list_as_xlsx(app_instance, meta_list, date_directory_path, save_path)

def table_to_meta_list(app_instance):
    """
    Convert table widget into list of dictionary for save as xlsx
    
    Returns:
        list: List of meta data dictionary
    """
    table = app_instance.iomanager_ui.table
    column_count = table.columnCount()
    row_count = table.rowCount()

    headers = [table.horizontalHeaderItem(col).text() for col in range(1, column_count)]
    meta_list = []

    for row in range(row_count):
        row_dict = {}
        for col in range(1, column_count):
            item = table.item(row, col)
            if item:
                text_value = item.text()
            # Handle QPixmap
            else:
                cell_widget = table.cellWidget(row, col)
                if cell_widget:
                    text_value = cell_widget.text()
                else:
                    text_value = ""
            row_dict[headers[col - 1]] = text_value

        meta_list.append(row_dict)
    return meta_list

