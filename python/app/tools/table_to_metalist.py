from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os

def save_table_to_xlsx(table, save_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    column_count = table.columnCount()
    row_count = table.rowCount()

    # checkbox 열(0) 스킵
    headers = []
    for col in range(1, column_count):
        headers.append(table.horizontalHeaderItem(col).text())
    ws.append(headers)

    for row in range(row_count):
        row_data = []
        for col in range(1, column_count):
            item = table.item(row, col)
            if item:
                value = item.text()
            else:
                value = ""
            row_data.append(value)

        ws.append(row_data)

        # 썸네일 처리
        if "thumbnail_path" in headers:
            try:
                thumb_col_idx = headers.index("thumbnail_path") + 1  # openpyxl 1부터 시작
                thumbnail_path = table.item(row, thumb_col_idx).text()
                if os.path.exists(thumbnail_path):
                    img_col_letter = get_column_letter(headers.index("thumbnail") + 1)
                    cell_ref = f"{img_col_letter}{row + 2}"  # +1(헤더) +1(1부터 시작)
                    img = Image(thumbnail_path)
                    img.width = 192
                    img.height = 108
                    ws.add_image(img, cell_ref)
            except Exception as e:
                print(f"[WARN] Thumbnail insert failed: {e}")

    wb.save(save_path)
    print(f"[OK] Table saved to {save_path}")
