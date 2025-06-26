import os
from openpyxl import load_workbook
import sgtk
import re

logger = sgtk.platform.get_logger(__name__)

def get_publish_list(app_instance, xl_path):
    if not os.path.exists(xl_path):
        return []
    
    wb = load_workbook(xl_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    keep_headers = headers[2:]
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        full_row = dict(zip(headers, row))

        if not full_row.get("check"):
            continue

        seq  = full_row.get("seq name")
        shot = full_row.get("shot name")
        if not seq or not shot:
            logger.info(f"Not enough information in excel file. Row {row_idx} skipped.")
            continue

        shot_dir = os.path.join(app_instance.project_path, "seq", seq, shot, "plate", "org")
        if not os.path.exists(shot_dir):
                versions = []
        else:
            versions = os.listdir(shot_dir)
        new_version = get_new_plate_version(versions)
        
        row_dict = {}
        for h in keep_headers:
            row_dict[h] = full_row.get(h)
        row_dict["version"] = new_version
        data.append(row_dict)    
    return data

def table_to_publish_list(app_instance):
    ui = app_instance.iomanager_ui
    table = ui.table
    data = []

    for row in range(table.rowCount()):
        cb = table.cellWidget(row, 0)
        checked = cb.isChecked()
        if not checked:
            continue

        row_dict = {}
        for col in range(2, table.columnCount()):
            header_item = table.horizontalHeaderItem(col)
            if not header_item:
                continue
            key = header_item.text()
            cell_item = table.item(row, col)
            if cell_item:
                row_dict[key] = cell_item.text()
            else:
                row_dict[key] = None

            seq  = row_dict.get("seq name")
            shot = row_dict.get("shot name")
            version = row_dict.get("version")
            if not seq or not shot or not version:
                logger.info(f"Not enough information in table. Row {row} skipped.")
                continue

            # shot_dir = os.path.join(app_instance.project_path, "seq", seq, shot, "plate", "org")
            # if not os.path.exists(shot_dir):
            #     versions = []
            # else:
            #     versions = os.listdir(shot_dir)
            # new_version = get_new_plate_version(versions)
            # row_dict["version"] = new_version
        data.append(row_dict)
    return data


def get_new_plate_version(dir_list):
    version_nums = []
    for name in dir_list:
        match = re.match(r"v(\d{3})", name)
        if match:
            version_nums.append(int(match.group(1)))
    if version_nums:
        latest_num = max(version_nums) + 1
        return f"v{latest_num:03d}"
    else:
        return "v001"

def upload_shotgrid(app_instance, paths, shot_info):
    if all(os.path.exists(path) for path in paths.values()):
        mp4_path = paths["mp4"]
        webm_path = paths["webm"]
        filmstrip_path = paths["montage"]
        thumbnail_path = paths["thumbnail"]

        seq = shot_info["seq name"]
        shot = shot_info["shot name"]
        scan_name = shot_info.get("scan name", "")
        scan_path = shot_info.get("scan path", "")
        org_path = shot_info.get("org path", "")
        resolution = shot_info.get("resolution", "")
        file_type = shot_info.get("file type", "")
        start_frame = shot_info.get("start frame", "")
        end_frame = shot_info.get("end frame", "")
        duration = shot_info.get("duration", "")
        fps = shot_info.get("fps", "")
        date_origin = shot_info.get("date", "")
        timecode_in = shot_info.get("timecode in", "")
        timecode_out = shot_info.get("timecode out", "")
        ver = shot_info["version"]

        project_id = app_instance.sg.find_one("Project", [["name", "is", app_instance.project_name]], ["id"])
        # Create Sequence entity
        sequence_name = seq
        sequence = app_instance.sg.find_one("Sequence", [
            ["project", "is", project_id],
            ["code", "is", sequence_name]
        ], ["code"])

        if not sequence:
            sequence_data = {
                "project": project_id,
                "code": sequence_name
            }
            sequence = app_instance.sg.create("Sequence", sequence_data)
            print(f"Sequence created: {sequence['code']}")
        

        # Create Shot entity
        shot_name = shot
        created_shot = app_instance.sg.find_one("Shot", [
            ["project", "is", project_id],
            ["code", "is", shot_name]
        ])

        if not created_shot:
            shot_data = {
                "project": project_id,
                "code": shot_name,
                "sg_sequence": sequence
            }
            created_shot = app_instance.sg.create("Shot", shot_data)
            print(f"Shot created: {created_shot['code']}")

            # Thumbnail upload to shot
            app_instance.sg.upload_thumbnail("Shot", created_shot["id"], thumbnail_path)
            print(f"Thumbnail uploaded sucessfully")

            # Flimstrip upload to shot
            app_instance.sg.upload_filmstrip_thumbnail("Shot", created_shot["id"], filmstrip_path)
            print(f"Filmstrip uploaded sucessfully")

        # Create version entity and link to shot
        version_name = f"{shot}_plate_{ver}"
        version_data = {
            "project": project_id,
            "code": version_name,
            "entity": created_shot,
            "description": f"Description for {shot}",
            "sg_scan_name": scan_name, 
            "sg_scan_path": scan_path,
            "sg_path_to_frames":org_path,
            "sg_resolution": resolution,
            "sg_ext": file_type,
            "sg_first_frame": int(start_frame),
            "sg_last_frame": int(end_frame),
            "frame_count": int(duration),
            "sg_date_origin": date_origin,
            "sg_timecode_in": timecode_in,
            "sg_timecode_out": timecode_out,
            "sg_uploaded_movie_frame_rate" : float(fps)
        }
        version = app_instance.sg.create("Version", version_data)
        print(f"Version created: {version['code']}")

        app_instance.sg.upload("Version", version["id"], mp4_path, field_name="sg_uploaded_movie")
        print(f"MP4 uploaded successfully")

        app_instance.sg.upload("Version", version["id"], webm_path, field_name="sg_uploaded_movie_webm")
        print(f"WEBM uploaded successfully")

        # jpg, filmstrip upload to Version
        app_instance.sg.upload_thumbnail("Version", version["id"], thumbnail_path)
        app_instance.sg.upload_filmstrip_thumbnail("Version", version["id"], filmstrip_path)
        logger.info(f"{shot_name} publish completed")
        return True
            
    else:
        app_instance.show_error_dialog("Publish failed.")
        for path in [mp4_path, webm_path, filmstrip_path, thumbnail_path]:
            error_msg = f"{shot_name} publish failed.\nMissing: {path}"
            logger.error(error_msg)
            if not os.path.exists(path):
                print(error_msg)
        return False