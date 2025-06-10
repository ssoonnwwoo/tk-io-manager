from tank.platform.qt import QtCore
for name, cls in QtCore.__dict__.items():
    if isinstance(cls, type): globals()[name] = cls

from tank.platform.qt import QtGui
for name, cls in QtGui.__dict__.items():
    if isinstance(cls, type): globals()[name] = cls

from ..controller.select_btn_clicked import on_select_clicked
from ..controller.table_checkbox_clicked import on_checkbox_clicked
from ..controller.save_btn_clicked import on_save_clicked
from ..controller.select_xl_btn_clicked import on_select_xl_clicked

from ..model.excel_manager import ExcelManager

from ..model.get_publish_info import get_publish_info
from ..model.rename import rename_sequence
from ..model.convert import exrs_to_jpgs, mov_to_exrs, exrs_to_video, exrs_to_montage, exrs_to_thumbnail
import os
import sgtk
import sys
from openpyxl import load_workbook

# Import resource binding python code
resource_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "resources"))
if resource_path not in sys.path:
    sys.path.append(resource_path)
import resources_rc

class IOManagerWidget(QWidget): 
    def __init__(self):
        super().__init__()
        # sgtk settings
        self.current_engine = sgtk.platform.current_engine()
        self.sg = self._get_shotgun_api()
        self.PROJECT_NAME = self._get_project_name()

        # path settings
        HOME_PATH = os.path.expanduser("~")
        # ex : HOME_PATH = /home/rapa
        self.DEFAULT_PATH = os.path.join(HOME_PATH, "show", self.PROJECT_NAME, "product", "scan") 
        # ex : DEFAULT_PATH = /home/rapa/show/{project_name}/product/scan

        # Create ExcelManager instance
        self.excel_manager = ExcelManager(self)

        self.checked_rows = []

        # Widgets Top
        p_label = QLabel("Current project: ")
        self.project_label = QLabel(self.PROJECT_NAME)
        file_path_label = QLabel("File path:")
        self.file_path_le = QLineEdit()
        self.file_path_le.setPlaceholderText("Input your shot path")
        self.file_path_le.setText(self.DEFAULT_PATH)
        shot_select_btn = QPushButton("Select")
        
        # Widgets Mid(Table)
        self.table = QTableWidget()
        
        # Widgets Bottom
        current_excel_label = QLabel("Currently displayed Excel file:")
        self.excel_label = QLabel("Ready to load")
        excel_save_btn = QPushButton("Version and Save")
        select_excel_btn = QPushButton("Select Excel")
        publish_btn = QPushButton("Publish")

        # Event Handle
        shot_select_btn.clicked.connect(lambda : on_select_clicked(self))
        excel_save_btn.clicked.connect(lambda : on_save_clicked(self))
        select_excel_btn.clicked.connect(lambda : on_select_xl_clicked(self))
        publish_btn.clicked.connect(self.on_publish_clicked)

        # Layout
        main_layout = QVBoxLayout()
        shot_select_container = QHBoxLayout()
        bottom_layout = QHBoxLayout()
        excel_group = QGroupBox("Excel")
        excel_btn_layout = QHBoxLayout()
        excel_btn_layout2 = QHBoxLayout()
        excel_container = QVBoxLayout()
        excel_label_conatainer = QVBoxLayout()
        
        shot_select_container.addWidget(p_label)
        shot_select_container.addWidget(self.project_label)
        shot_select_container.addStretch()
        shot_select_container.addWidget(file_path_label)
        shot_select_container.addWidget(self.file_path_le)
        shot_select_container.addWidget(shot_select_btn)
        #shot_select_container.addWidget(shot_load_btn)
        
        excel_label_conatainer.addWidget(current_excel_label, alignment=Qt.AlignTop)
        excel_label_conatainer.addWidget(self.excel_label, alignment=Qt.AlignTop)
        excel_btn_layout.addWidget(excel_save_btn)
        excel_btn_layout2.addWidget(select_excel_btn)
        excel_container.addLayout(excel_btn_layout)
        excel_container.addLayout(excel_btn_layout2)
        excel_group.setLayout(excel_container)
        bottom_layout.addLayout(excel_label_conatainer)
        bottom_layout.addWidget(excel_group)
        bottom_layout.addWidget(publish_btn)

        main_layout.addLayout(shot_select_container)
        main_layout.addWidget(self.table)
        main_layout.addLayout(bottom_layout)

        self.setLayout(main_layout)



    def _get_project_name(self):
        """
        Get the project name from engine & context that is currently running

        Returns: 
            str: project name
        """
        context = self.current_engine.context
        project_name = context.project.get("name")
        return project_name

    def _get_shotgun_api(self):
        """
        Get Shotgun API from engine & sgtk

        Returns: 
            object: shotgun api
        """
        # Grab the already created Sgtk instance from the current engine.
        tk = self.current_engine.sgtk
        sg = tk.shotgun
        return sg

    def update_table(self, xlsx_path):
        """
        Update table widget based on excel file

        Args:
            str: Path of reference excel file
        """
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Making headers : First row of xlsx file -> Header label of table widget 
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(cell.value)
        header_list = ["check"] + headers

        self.table.setColumnCount(len(header_list))
        data_rows = list(ws.iter_rows(min_row=2, values_only=False))
        self.table.setRowCount(len(data_rows))
        self.table.setHorizontalHeaderLabels(header_list)

        for row_idx, row_data in enumerate(data_rows):
            checkbox = QCheckBox()
            checkbox.clicked.connect(lambda _, row=row_idx: on_checkbox_clicked(self, row, xlsx_path))
            self.table.setCellWidget(row_idx, 0, checkbox)

            for col_idx, cell in enumerate(row_data):
                header = headers[col_idx]
                value = cell.value
                if value == None:
                    value = ""

                if header == "thumbnail":
                    self.set_thumbnail_cell(row_idx, col_idx + 1, value)
                else:
                    item = QTableWidgetItem(str(value))
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                    self.table.setItem(row_idx, col_idx + 1, item)

    def set_thumbnail_cell(self, row, col, img_path):
        """
        Set thumbnail cell of table widget using QPixmap

        Args:
            row(int)
            col(int)
            img_path(str)
        """
        label = QLabel()
        pixmap = QPixmap(img_path)
        if not pixmap.isNull():
            label.setPixmap(pixmap.scaledToWidth(400, Qt.SmoothTransformation))
            label.setAlignment(Qt.AlignLeft)
            self.table.setCellWidget(row, col, label)
            self.table.setRowHeight(row, 230)
            self.table.setColumnWidth(col, 400)

    def get_checked_rows(self):
        checked_rows = []
        for row in range(self.table.rowCount()):
            checkbox = self.table.cellWidget(row, 0)
            if isinstance(checkbox, QCheckBox) and checkbox.isChecked():
                checked_rows.append(row + 2)
        return checked_rows

    def on_publish_clicked(self):
        xlsx_file_path = self.excel_label.text()
        checked_rows = self.get_checked_rows()
        if not checked_rows:
            QMessageBox.information(self, "No Selection", "Please check at least one item to publish.", QMessageBox.Ok)
            return
        shot_info_list = get_publish_info(xlsx_file_path, checked_rows)
        home_dir = os.path.expanduser("~")
        base_path = os.path.join(home_dir, "show")
        project_name = self.project_label.text()
        for shot_info in shot_info_list:
            seq = shot_info["sequence"]
            shot = shot_info["shot"]
            typ = shot_info["type"]
            ver = shot_info["version"]
            ver_jpg = f"{ver}_jpg"
            directory = shot_info["directory"]

            print(f"[INFO] Preparing to publish shot : {shot}")
            org_path = os.path.join(
            base_path, project_name,"seq",
            seq,shot, "plate", typ, ver
            )

            jpg_path = os.path.join(
            base_path, project_name,"seq",
            seq, shot, "plate", typ, ver_jpg
            )

            sg_path = os.path.join(
                base_path, project_name, "seq",
                seq, shot, "plate", "shotgrid_upload_datas", ver
            )

            os.makedirs(org_path, exist_ok=True)
            os.makedirs(jpg_path, exist_ok=True)
            os.makedirs(sg_path, exist_ok=True)

            files = []
            for file in os.listdir(directory):
                # full_path = os.path.join(directory, file)
                files.append(file)

            # 확장자 조사
            exts = set()
            for file in files:
                _, ext = os.path.splitext(file)
                exts.add(ext.lower())
            
            # exr -> rename + exr sequence to jpg sequence
            if ".exr" in exts:
                rename_sequence(directory, org_path)
                exrs_to_jpgs(org_path, jpg_path)
                mp4_path = exrs_to_video(org_path, sg_path, vformat='mp4')
                webm_path = exrs_to_video(org_path, sg_path, vformat='webm')
                filmstrip_path = exrs_to_montage(org_path, sg_path)
                thumbnail_path = exrs_to_thumbnail(org_path, sg_path)
            
            # mov -> mov to exr sequence + exr sequence to jpg sequence
            elif ".mov" in exts:
                mov_path = os.path.join(directory,files[0])
                success = mov_to_exrs(mov_path, org_path)
                if success:
                    exrs_to_jpgs(org_path, jpg_path)
                    mp4_path = exrs_to_video(org_path, sg_path, vformat='mp4')
                    webm_path = exrs_to_video(org_path, sg_path, vformat='webm')
                    filmstrip_path = exrs_to_montage(org_path, sg_path)
                    thumbnail_path = exrs_to_thumbnail(org_path, sg_path)

            if all(os.path.isfile(path) for path in [mp4_path, webm_path, filmstrip_path, thumbnail_path]):
                project_id = self.sg.find_one("Project", [["name", "is", project_name]], ["id"])
                # Create sequence
                # 시퀀스 존재 여부 확인
                sequence_name = seq
                sequence = self.sg.find_one("Sequence", [
                    ["project", "is", project_id],
                    ["code", "is", sequence_name]
                ], ['code'])

                # 없으면 생성
                if not sequence:
                    sequence_data = {
                        "project": project_id,
                        "code": sequence_name
                    }
                    sequence = self.sg.create("Sequence", sequence_data)
                    print(f"[COMPLETE] Sequence created: {sequence['code']}")
                else:
                    print(f"[INFO] Sequence already exists: {sequence['code']}")

                # Create shot
                shot_name = f"{shot}_{ver}"
                shot_data = {
                    "project": project_id,
                    "code": shot_name,          
                    "sg_sequence": sequence
                }
                created_shot = self.sg.create("Shot", shot_data)
                print(f"[COMPLETE] Created Shot: {created_shot['code']}")

                # thumbnail upload to shot
                self.sg.upload_thumbnail("Shot", created_shot["id"], thumbnail_path)
                print(f"[COMPLETE] Uploaded thumbnail: {thumbnail_path}")

                # filmstrip upload to shot
                self.sg.upload_filmstrip_thumbnail("Shot", created_shot["id"], filmstrip_path)
                print(f"[COMPLETE] Uploaded filmstrip thumbnail: {filmstrip_path}")

                # create version and link to shot
                version_name = f"{shot}_plate_{ver}"
                version_data = {
                    "project": project_id,
                    "code": version_name,
                    "entity": created_shot,
                    "description": f"Description for {shot}"
                }
                version = self.sg.create("Version", version_data)
                print(f"[COMPLETE] Created Version: {version['code']}")
                
                # mp4 upload to version
                self.sg.upload("Version", version["id"], mp4_path, field_name="sg_uploaded_movie")
                print(f"[COMPLETE] Uploaded mp4 movie: {mp4_path}")

                # jpg, filmstrip upload to Version
                self.sg.upload_thumbnail("Version", version["id"], thumbnail_path)
                self.sg.upload_filmstrip_thumbnail("Version", version["id"], filmstrip_path)

                print(f"[COMPLETE]: {shot} publish completed\n\n")

            else:
                print("[ERROR] One or more expected output files are missing:")
                for path in [mp4_path, webm_path, filmstrip_path, thumbnail_path]:
                    if not os.path.isfile(path):
                        print(f" - Missing: {path}")