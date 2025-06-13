from openpyxl import load_workbook
from sgtk.platform.qt import QtGui

def update_table(app_instance):
    """
    Update table widget based on excel file

    Args:
        str: Path of reference excel file
    """
    xlsx_path = app_instance.xl_manager.get_xl_path()
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Making headers : First row of xlsx file -> Header label of table widget 
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)
    # header_list = headers

    app_instance.iomanager_ui.table.setColumnCount(len(headers))
    data_rows = list(ws.iter_rows(min_row=2, values_only=False))
    app_instance.iomanager_ui.table.setRowCount(len(data_rows))
    app_instance.iomanager_ui.table.setHorizontalHeaderLabels(headers)

    check_idx = headers.index("check")

    for row_idx, row_data in enumerate(data_rows):
        checkbox = QtGui.QCheckBox()
        # checkbox.clicked.connect(lambda _, row=row_idx: on_checkbox_clicked(app_instance, row, xlsx_path))
        app_instance.iomanager_ui.table.setCellWidget(row_idx, check_idx, checkbox)

        for col_idx, cell in enumerate(row_data):
            header = headers[col_idx]
            value = cell.value
            if value == None:
                value = ""

            if header == "thumbnail":
                set_thumbnail_cell(app_instance, row_idx, col_idx, value)
            elif header == "check":
                set_checkbox(app_instance, row_idx, col_idx, value)
            else:
                item = QtGui.QTableWidgetItem(str(value))
                item.setFlags(item.flags() | QtGui.Qt.ItemIsEditable)
                app_instance.iomanager_ui.table.setItem(row_idx, col_idx, item)

# 수정 예정
def on_checkbox_clicked(app_instance, row, xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    seq_idx = headers.index("seq name")
    shot_idx = headers.index("shot name")

    excel_row = row + 2
    seq = ws.cell(row=excel_row, column=seq_idx + 1).value
    shot = ws.cell(row=excel_row, column=shot_idx + 1).value

    if not seq or not shot or str(seq).strip() == "" or str(shot).strip() == "":
        QtGui.QMessageBox.warning(
            app_instance.iomanager_ui,
            "Missing Data",
            f"Row[{row + 1}] missing 'seq name' or 'shot name'.\nThis row cannot be selected.",
            QtGui.QMessageBox.Ok
        )

        checkbox = app_instance.iomanager_ui.table.cellWidget(row, 0)
        if isinstance(checkbox, QtGui.QCheckBox):
            checkbox.setChecked(False)

def set_thumbnail_cell(app_instance, row, col, img_path):
    """
    Set thumbnail cell of table widget using QPixmap

    Args:
        row(int)
        col(int)
        img_path(str)
    """
    label = QtGui.QLabel()
    pixmap = QtGui.QPixmap(img_path)
    if not pixmap.isNull():
        label.setPixmap(pixmap.scaledToWidth(400, QtGui.Qt.SmoothTransformation))
        label.setAlignment(QtGui.Qt.AlignLeft)
        app_instance.iomanager_ui.table.setCellWidget(row, col, label)
        app_instance.iomanager_ui.table.setRowHeight(row, 230)
        app_instance.iomanager_ui.table.setColumnWidth(col, 400)

def set_checkbox(app_instance, row, col, checked):
    """
    Set checkbox of table widget to checked or unchecked based on value

    Args:
        app_instance: Main dialog or app object
        row (int): Table row index
        col (int): Table column index (where checkbox is)
        checked (int ): 0, 1
    """
    cell_widget = app_instance.iomanager_ui.table.cellWidget(row, col)
    if isinstance(cell_widget, QtGui.QCheckBox):
        if checked == 1:
            cell_widget.setChecked(True)
        else:
            cell_widget.setChecked(False)
