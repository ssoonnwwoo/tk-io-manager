from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
import re
import pyseq
import subprocess
import json
from tank.platform.qt import QtGui

class ExcelManager:
    def __init__(self, iomanager_instance):
        self.date_directory_path = ""
        self.iomanager = iomanager_instance
    
    def set_date_path(self, date_directory_path):
        self.date_directory_path = date_directory_path

    def show_error_dialog(self, message):
        QtGui.QMessageBox.warning(
            self.iomanager,
            "Error",
            message,
            QtGui.QMessageBox.Ok
        )

    def get_xl_files(self):
        versioned_files = []
        if not self.date_directory_path:
            self.show_error_dialog(f"Please select date directory first")
            return versioned_files
        # ex ) date_prefix: 20250529
        date_prefix = os.path.basename(self.date_directory_path)
        pattern = re.compile(rf"{date_prefix}_list_v(\d{{3}})\.xlsx$")
        for fname in os.listdir(self.date_directory_path):
            match = pattern.match(fname)
            if match:
                version = int(match.group(1))
                versioned_files.append((version, fname))
        # ex ) versioned_files: [(1, 20250529_list_v001.xlsx), (2, 20250529_list_v002.xlsx), ..., (14, 20250529_list_v014.xlsx)]
        return versioned_files

    def get_latest_xlsx(self):
        meta_data_list = self.export_metadata(self.date_directory_path)
        versioned_files = self.get_xl_files()
        if not versioned_files:
            # If there is no xlsx file. Make _v001.xlsx
            print("[INFO] No versioned file found. v001 will be created")
            new_xlsx_path = self.save_list_as_xlsx(meta_data_list)
            latest_xlsx_path = new_xlsx_path
        else:
            # return latest version filename
            latest_version_tuple = max(versioned_files)
            latest_xlsx_path = os.path.join(self.date_directory_path, latest_version_tuple[1])
        return latest_xlsx_path
    
    def get_new_xl(self):
        """
        Get (latest_version + 1) xlsx file path
        """
        versioned_files = self.get_xl_files()
        # No selected date path 
        if not versioned_files:
            # self.show_error_dialog(f"Please select date directory first")
            new_version_xl = ""
            return new_version_xl
        latest_version_tuple = max(versioned_files)
        next_version = latest_version_tuple[0] + 1
        file_name = f"{self.date_directory_path}_list_v{next_version:03d}.xlsx"
        new_version_xl = os.path.join(self.date_directory_path, file_name)
        return new_version_xl
    

    def export_metadata(self, date_path):
        # meta data list: one of xlsx rows 
        meta_list = []
        thumbnails_dir = os.path.join(date_path, "thumbnails")
        os.makedirs(thumbnails_dir, exist_ok=True)

        # .../scan/{date} 순회 하면서 Sequence 객체 get
        for root, dirs, seqs in pyseq.walk(date_path):
            dirs[:] = [d for d in dirs if d != "thumbnails"]
            first_frame_path = "" #exr의 첫 프레임 path / mov path
            for seq in seqs:
                _, ext = os.path.splitext(seq.name)
                # EXR sequnece 폴더의 첫 프레임을 thumnail(JPG)로 변환
                if ext == ".exr":
                    first_frame_path = seq[0].path
                    thumb_name = seq.head().strip(".") + ".jpg"
                    thumb_path = os.path.join(thumbnails_dir, thumb_name)
                    # If thumbnail not exist, make thumbnail
                    if not os.path.exists(thumb_path):
                        make_excel_thumbnail_result = self.make_excel_thumbnail(first_frame_path, thumb_path)
                        if not make_excel_thumbnail_result:
                            QtGui.QMessageBox.warning(
                                self.iomanager,
                                "Error",
                                f"Error occurred while attempting to convert thumbnail.\nIO Manager skips the thumbnail creation process.",
                                QtGui.QMessageBox.Ok
                            )
                            thumb_path = ""
                    # Skip making thumbnail if thumbnail exist
                    else:
                        print(f"[SKIP] Thumbnail already exist: {thumb_path}")
                # MOV 폴더의 첫 프레임 thumbnial(JPG) 추출
                elif ext == ".mov":
                    # mov를 shot 단위의 exr로 만들고 이후 과정은 똑같이
                    # mov_path = os.path.join(root, seq.name)
                    # scan_data_path = mov_path # for export meta data
                    # thumb_name = os.path.splitext(seq.name)[0] + ".jpg"
                    # thumb_path = os.path.join(thumbnails_dir, thumb_name)
                    # if not os.path.exists(thumb_path):
                    #     if mov_to_jpg(scan_data_path, thumb_path):
                    #         print(f"[OK] Thumbnail created: {thumb_path}")
                    #     else:
                    #         print(f"[FAIL] Thumbnail create failed: {scan_data_path}")
                    #         thumb_path = ""

                    # else:
                    #     print(f"[SKIP] Thumbnail already exist: {thumb_path}")
                    continue
                else:
                    print(f"[SKIP] {seq} is not EXR OR MOV")
                    continue

                # Exporting metadata using EXIF tool 
                result = subprocess.run(
                    ["exiftool", "-json", first_frame_path],
                    capture_output=True, 
                    text=True
                )
                # meta data json 파싱
                # ex) result(string) -> meta(json) 
                meta = json.loads(result.stdout)[0]
                meta["thumbnail"] = thumb_path
                meta_list.append(meta)
        return meta_list

    def make_excel_thumbnail(self, input_exr_path, output_jpg_path):        
        cmd = [
            "oiiotool",
            input_exr_path,
            "--colorconvert", "ACES", "sRGB",
            "-o", output_jpg_path
        ]

        result = subprocess.run(cmd, check=False)
        if result.returncode != 0:
            print(f"[ERROR] Error occurred while attempting to convert thumbnail")
            return False
        print(f"[COMPLETE] {input_exr_path} >>>>> {output_jpg_path}")
        return True

    def save_list_as_xlsx(self, meta_data_list, xl_name = None):
        """
        Save meta datas as xlsx using openpyxl

        meta_data : data, one row of xlsx rows
        meta_data_list = [meta_data1, meta_data2, ... , meta_dataN]
        """
        if not meta_data_list:
            QtGui.QMessageBox.warning(self.iomanager, "No shot error", "No shot for exporting excel file")
            return None
        if not xl_name:
            xl_name = f"{os.path.basename(self.date_directory_path)}_list_v001.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Metadata"

        default_fields = ["thumbnail", "shot_name", "seq_name"]
        all_keys_set = set()
        for meta_data in meta_data_list:
            for key in meta_data.keys():
                all_keys_set.add(key)

        all_keys_set -= set(default_fields)

        all_fields = default_fields + list(all_keys_set) # headers
        ws.append(all_fields)

        # Insert meta data to work sheet
        for row_idx, meta_data in enumerate(meta_data_list, start=2): # Start at row2(Row1 : header)
            for col_idx, field in enumerate(all_fields, start=1):
                value = meta_data.get(field, "")
                # if value's type is list
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                ws.cell(row=row_idx, column=col_idx, value=value)

            # Insert thumbnail to cell
            thumb_path = meta_data.get("thumbnail")
            if thumb_path and os.path.exists(thumb_path):
                col_letter = get_column_letter(all_fields.index("thumbnail") + 1)
                cell_ref = f"{col_letter}{row_idx}"
                img = Image(thumb_path)
                img.width = 192
                img.height = 108
                #ws[cell_ref].value = None
                ws.add_image(img, cell_ref)
        xl_path = os.path.join(self.date_directory_path, xl_name)
        wb.save(xl_path)

        print(f"[COMPLETE] Metadata exported to:\n{xl_path}")
        return xl_path
    
    def save_table_to_xlsx(self, save_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Metadata"

        table = self.iomanager.table

        column_count = table.columnCount()
        row_count = table.rowCount()

        # checkbox 열(0) 스킵
        headers = []
        for col in range(1, column_count):
            headers.append(table.horizontalHeaderItem(col).text())
        ws.append(headers)
        
        if "thumbnail" in headers:
            thumb_col_idx = headers.index("thumbnail") + 1

        for row in range(row_count):
            row_data = []
            for col in range(1, column_count):
                print(col)
                if thumb_col_idx == col:
                    print(f"Catchyou {col}")
                    value = table.cellWidget(row, col + 4)
                    print(value.text())
                if col != thumb_col_idx:
                    if not value:
                        value = ""
                else:
                    pass
                    # print(table.item(row, col))
                    # print(table.item(row, col))
                row_data.append(value)

            ws.append(row_data)

            # 썸네일 처리
            # if "thumbnail" in headers:
            #     thumb_col_idx = headers.index("thumbnail") + 1  # openpyxl 1부터 시작
            #     print(table.item(row, thumb_col_idx))
        #         try:
        #             thumb_col_idx = headers.index("thumbnail") + 1  # openpyxl 1부터 시작
        #             thumbnail_path = table.item(row, thumb_col_idx).text()
        #             print(thumbnail_path)
        #             if os.path.exists(thumbnail_path):
        #                 img_col_letter = get_column_letter(headers.index("thumbnail") + 1)
        #                 cell_ref = f"{img_col_letter}{row + 2}"  # +1(헤더) +1(1부터 시작)
        #                 img = Image(thumbnail_path)
        #                 img.width = 192
        #                 img.height = 108
        #                 ws.add_image(img, cell_ref)
        #         except Exception as e:
        #             print(f"[WARN] Thumbnail insert failed: {e}")

        # wb.save(save_path)
        # print(f"[OK] Table saved to {save_path}")