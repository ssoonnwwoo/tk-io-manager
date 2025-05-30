from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
from PySide6.QtWidgets import QMessageBox

def save_as_xlsx(date_path, file_name, meta_list):
    if not meta_list:
            QMessageBox.warning(None, "Load Stopped", "No shots for exporting excel file")
            return None
    if not file_name:
        file_name = f"{os.path.basename(os.path.normpath(date_path))}_list_v001.xlsx"
    # dictionary로 이루어진 list -> xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    default_fields = ["thumbnail", "thumbnail_path", "shot_name", "seq_name"]
    all_keys_set = set()
    for m in meta_list:
        for key in m.keys():
            all_keys_set.add(key)

    all_fields = default_fields + list(all_keys_set)
    
    # 헤더 작성
    ws.append(all_fields)

    # 데이터 작성
    for row_idx, meta in enumerate(meta_list, start=2):  # 2부터 시작 (1행은 헤더)
        for col_idx, field in enumerate(all_fields, start=1): # 1부터 시작
            value = meta.get(field, "")
            # list인 meta data -> str 처리
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            ws.cell(row=row_idx, column=col_idx, value=value)

        thumb_path = meta.get("thumbnail_path")
        if thumb_path and os.path.exists(thumb_path):
            col_letter = get_column_letter(all_fields.index("thumbnail") + 1)
            cell_ref = f"{col_letter}{row_idx}"
            try:
                img = Image(thumb_path)
                img.width = 192  
                img.height = 108
                ws.add_image(img, cell_ref)
            except Exception as e:
                print(f"[WARN] Thumbnail insert failed: {e}")

    xlsx_path = os.path.join(date_path, file_name)

    wb.save(xlsx_path)

    # QMessageBox.information(None, "Export Complete", f"Metadata exported to:\n{xlsx_path}")
    print(f"[COMPLETE] Metadata exported to:\n{xlsx_path}")
    return xlsx_path