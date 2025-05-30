from openpyxl import load_workbook

def extract_directory_column(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    dir_col_idx = headers.index("Directory")

    # 2행부터 Directory 열 데이터 추출
    directories = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # print(row)
        directories.append(row[dir_col_idx])

    return directories
