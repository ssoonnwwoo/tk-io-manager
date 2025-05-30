from openpyxl import load_workbook
import os
import re

def get_latest_plate_version(dir_list):
    version_nums = []
    for name in dir_list:
        match = re.match(r"v(\d{3})", name)
        if match:
            version_nums.append(int(match.group(1)))
    if version_nums:
        latest_num = max(version_nums) + 1
        return f"v{latest_num:03d}"
    else:
        return "v001"

def get_publish_info(xlsx_file_path, checked_rows):
    if os.path.exists(xlsx_file_path):
        wb = load_workbook(xlsx_file_path)
        ws = wb.active

        parts = os.path.normpath(xlsx_file_path).split(os.sep)
        show_idx = parts.index("show")
        project_name = parts[show_idx + 1]

        # 첫 번째 행에서 헤더 인덱스 찾기
        # headers ex : {'seq_name': 0, 'shot_name': 1, 'type': 2, 'Directory': 3}
        headers = {}
        for col_idx, cell in enumerate(ws[1]):
            headers[cell.value] = col_idx

        # 각 행에서 dict 구성
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row_idx not in checked_rows:
                continue

            seq = row[headers["seq_name"]]
            shot = row[headers["shot_name"]]
            #typ = row[headers["type"]]
            directory = row[headers["Directory"]]

            if not (seq and shot and directory):
                print(f"[SKIP] Not enough information in excel file : row num {row_idx}")
                continue  # 불완전한 행은 일단 건너뜀
            
            # Get latest version number
            home_dir = os.path.expanduser("~")
            base_dir = os.path.join(home_dir, "show")
            #base_dir = "/home/rapa/show"
            shot_dir = os.path.join(base_dir, project_name, "seq", seq, shot, "plate/org")
            try:
                versions = os.listdir(shot_dir)
            except FileNotFoundError:
                versions = []
            latest_version = get_latest_plate_version(versions)
            
            data.append({
                "sequence": str(seq),
                "shot": str(shot),
                "type": "org",
                "version": latest_version,
                "directory" : str(directory)
            })
        return data
    else:
        print(f"Not valid path : {xlsx_file_path} ")
        return
    
    # .../show/{project}/seq/{sequence}/{shot}/plate/org/

