# Copyright (c) 2013 Shotgun Software Inc.
#
# CONFIDENTIAL AND PROPRIETARY
#
# This work is provided "AS IS" and subject to the Shotgun Pipeline Toolkit
# Source Code License included in this distribution package. See LICENSE.
# By accessing, using, copying or modifying this work you indicate your
# agreement to the Shotgun Pipeline Toolkit Source Code License. All rights
# not expressly granted therein are reserved by Shotgun Software Inc.

import sgtk
import os
from openpyxl import load_workbook
from .get_latest_xl import get_latest_xlsx
from .xl_to_table import update_table
from sgtk.platform.qt import QtGui
from .model.excel_manager import ExcelManager
from .view.iomanager_ui import IOManagerWidget

# standard toolkit logger
logger = sgtk.platform.get_logger(__name__)


def show_dialog(app_instance):
    """
    Shows the main dialog window.
    """
    app_instance.engine.show_dialog("IO Manager", app_instance, AppDialog)


class AppDialog(QtGui.QWidget):
    """
    Main application dialog window
    """

    def __init__(self):
        """
        Constructor
        """
        QtGui.QWidget.__init__(self)
        self.current_engine = sgtk.platform.current_engine()
        self.sg = self.get_shotgun_api()
        self.project_name = self.get_project_name()
        self.home_path = os.path.expanduser("~")
        self.project_path = os.path.join(self.home_path, "show", self.project_name)
        self.default_path = os.path.join(self.project_path, "product", "scan")
        # ex : default_path = {self.home_path}/show/{project_name}/product/scan
        
        # View
        self.iomanager_ui = IOManagerWidget()
        self.connect_event_handlers()
        self.iomanager_ui.project_label.setText(self.project_name)
        
        # Model
        self.xl_manager = ExcelManager()

        layout = QtGui.QVBoxLayout()
        layout.addWidget(self.iomanager_ui)
        self.setLayout(layout)
        self.resize(1400, 800)


        logger.info("Launching IO Manager Dev...")


    def connect_event_handlers(self):
        self.iomanager_ui.shot_select_btn.clicked.connect(self.on_select_clicked)

    def on_select_clicked(self):
        selected_date_path = self.select_date_directory()
        if not selected_date_path:
            return
        self.xl_manager.set_date_path(selected_date_path)
        self.iomanager_ui.file_path_le.setText(selected_date_path)
        latest_xlsx_path = get_latest_xlsx(self)
        self.xl_manager.set_xl_path(latest_xlsx_path)
        update_table(self)
        self.iomanager_ui.excel_label.setText(latest_xlsx_path)

    def on_checkbox_clicked(self, row, xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        seq_idx = headers.index("seq name")
        shot_idx = headers.index("shot name")

        excel_row = row + 2
        seq = ws.cell(row=excel_row, column=seq_idx + 1).value
        shot = ws.cell(row=excel_row, column=shot_idx + 1).value

        if not seq or not shot or str(seq).strip() == "" or str(shot).strip() == "":
            QtGui.QMessageBox.warning(
                self.iomanager_ui,
                "Missing Data",
                f"Row[{row + 1}] missing 'seq name' or 'shot name'.\nThis row cannot be selected.",
                QtGui.QMessageBox.Ok
            )

            checkbox = self.iomanager_ui.table.cellWidget(row, 0)
            if isinstance(checkbox, QtGui.QCheckBox):
                checkbox.setChecked(False)


    def select_date_directory(self):
        default_path = self.default_path
        if not os.path.isdir(default_path):
            default_path = self.project_path
        
        selected_date_path = QtGui.QFileDialog.getExistingDirectory(
            self.iomanager_ui, # Parent Widget
            "Select scan data folder",
            default_path, 
            QtGui.QFileDialog.ShowDirsOnly
        )

        # return date directory path
        # ex) date directory path : /home/rapa/show/my_project/product/scan/250529
        if selected_date_path:
            if not selected_date_path.startswith(self.project_path):
                QtGui.QMessageBox.warning(
                    self.iomanager_ui, 
                    "Warning", 
                    f"'{selected_date_path}' Out of boundary : Not Selected Project"
                )
                return None
            return selected_date_path
        return None

    def get_project_name(self):
        """
        Get the project name from engine & context that is currently running

        Returns: 
            str: project name
        """
        context = self.current_engine.context
        project_name = context.project.get("name")
        return project_name

    def get_shotgun_api(self):
        """
        Get Shotgun API from engine & sgtk

        Returns: 
            object: shotgun api
        """
        # Grab the already created Sgtk instance from the current engine.
        tk = self.current_engine.sgtk
        sg = tk.shotgun
        return sg
    
    def show_error_dialog(self, message):
        QtGui.QMessageBox.warning(
            self.iomanager,
            "Error",
            message,
            QtGui.QMessageBox.Ok
        )